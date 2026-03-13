"""Bulk transaction import from workbook files."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import cast
from uuid import UUID

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell import Cell  # type: ignore[import-untyped]
from openpyxl.utils.datetime import from_excel  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions import AccountNotFoundError
from app.models.types import TransactionType
from app.repositories.account_repository import AccountRepository
from app.repositories.transaction_repository import TransactionRepository
from app.schemas.finance import TransactionImportError


@dataclass
class ParsedImportRow:
    """Normalized transaction row parsed from a workbook."""

    row_number: int
    amount: Decimal
    type: TransactionType
    description: str
    date_accrual: datetime
    date_cash: datetime


@dataclass
class ImportRowParseError(ValueError):
    """Structured per-row parsing error."""

    code: str
    message: str

    def __str__(self) -> str:
        return self.message


class TransactionImportService:
    """Service for importing actual transactions from XLSX workbooks."""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.account_repo = AccountRepository(session)
        self.transaction_repo = TransactionRepository(session)

    async def import_workbook(
        self,
        *,
        user_id: UUID,
        account_id: UUID,
        workbook_bytes: bytes,
        filename: str,
    ) -> tuple[list[UUID], list[TransactionImportError]]:
        """Import workbook rows into actual transactions for a specific account."""
        if not filename.lower().endswith(".xlsx"):
            raise ValueError("Only .xlsx files are supported")

        account = await self.account_repo.get_by_id(account_id)
        if account is None or account.user_id != user_id:
            raise AccountNotFoundError(str(account_id))

        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
        worksheet = workbook.worksheets[0]

        parsed_rows: list[ParsedImportRow] = []
        errors: list[TransactionImportError] = []

        for row_number, row in enumerate(worksheet.iter_rows(), start=1):
            if self._is_empty_row(row):
                continue

            try:
                parsed_rows.append(self._parse_row(row_number=row_number, row=row, epoch=workbook.epoch))
            except ImportRowParseError as exc:
                errors.append(
                    TransactionImportError(
                        row_number=row_number,
                        code=exc.code,
                        message=str(exc),
                    )
                )

        imported_ids: list[UUID] = []
        for row in parsed_rows:
            transaction = await self.transaction_repo.create(
                user_id=user_id,
                account_id=account_id,
                amount=row.amount,
                type_=row.type,
                date_accrual=row.date_accrual,
                date_cash=row.date_cash,
                description=row.description,
            )
            imported_ids.append(transaction.id)

        return imported_ids, errors

    def _parse_row(
        self,
        *,
        row_number: int,
        row: tuple[Cell, ...],
        epoch: object,
    ) -> ParsedImportRow:
        if len(row) < 3:
            raise ImportRowParseError(
                code="import_row_missing_columns",
                message="Expected at least 3 columns: date, description, amount",
            )

        date_value = self._parse_date_cell(row[0], epoch)
        description = self._parse_description_cell(row[1])
        amount, transaction_type = self._parse_amount_cell(row[2])

        timestamp = self._to_utc_datetime(date_value)
        return ParsedImportRow(
            row_number=row_number,
            amount=amount,
            type=transaction_type,
            description=description,
            date_accrual=timestamp,
            date_cash=timestamp,
        )

    def _is_empty_row(self, row: tuple[Cell, ...]) -> bool:
        return all(cell.value in (None, "") for cell in row[:3])

    def _parse_date_cell(self, cell: Cell, epoch: object) -> date | datetime:
        value = cell.value
        if value is None or value == "":
            raise ImportRowParseError(code="import_row_missing_date", message="Missing date")

        if isinstance(value, datetime):
            return value

        if isinstance(value, date):
            return value

        if isinstance(value, (int, float)):
            return cast(date | datetime, from_excel(value, epoch=epoch))

        if isinstance(value, str):
            cleaned = value.strip()
            for parser in (
                "%Y-%m-%d",
                "%d.%m.%Y",
                "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S",
            ):
                try:
                    return datetime.strptime(cleaned, parser)
                except ValueError:
                    continue
            raise ImportRowParseError(
                code="import_row_invalid_date",
                message="Unsupported date format",
            )

        raise ImportRowParseError(
            code="import_row_invalid_date",
            message="Unsupported date value",
        )

    def _parse_description_cell(self, cell: Cell) -> str:
        value = cell.value
        if value is None:
            raise ImportRowParseError(
                code="import_row_missing_description",
                message="Missing description",
            )

        description = str(value).strip()
        if not description:
            raise ImportRowParseError(
                code="import_row_missing_description",
                message="Missing description",
            )

        return description

    def _parse_amount_cell(self, cell: Cell) -> tuple[Decimal, TransactionType]:
        value = cell.value
        if value is None or value == "":
            raise ImportRowParseError(code="import_row_missing_amount", message="Missing amount")

        if isinstance(value, (int, float, Decimal)):
            decimal_value = Decimal(str(value))
        elif isinstance(value, str):
            normalized = (
                value.replace("\u00a0", "")
                .replace(" ", "")
                .replace("₽", "")
                .replace("$", "")
                .replace(",", ".")
            )
            try:
                decimal_value = Decimal(normalized)
            except InvalidOperation as exc:
                raise ImportRowParseError(
                    code="import_row_invalid_amount",
                    message="Unsupported amount format",
                ) from exc
        else:
            raise ImportRowParseError(
                code="import_row_invalid_amount",
                message="Unsupported amount value",
            )

        if decimal_value == 0:
            raise ImportRowParseError(
                code="import_row_zero_amount",
                message="Amount cannot be zero",
            )

        transaction_type = (
            TransactionType.INCOME if decimal_value > 0 else TransactionType.EXPENSE
        )
        return abs(decimal_value), transaction_type

    def _to_utc_datetime(self, value: date | datetime) -> datetime:
        if isinstance(value, datetime):
            if value.tzinfo is not None:
                return value.astimezone(UTC)
            return value.replace(tzinfo=UTC)

        return datetime.combine(value, time.min, tzinfo=UTC)
